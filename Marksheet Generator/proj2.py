import csv
import pandas as pd
import math
import openpyxl 
import os
from pywebio import *
from pywebio.input import *
from pywebio.output import *
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font, colors,Border,Side
from openpyxl.drawing.image import Image
import smtplib, ssl
from email.message import EmailMessage

class QuizMarksheetGenerator:
    def __init__(self):
        self.roll_to_email = {}

    def send_individual_mail(self, server, file, email):
        msg = EmailMessage()
        msg['Subject'] = "Quiz marks CSE"
        msg['From'] = "Sahil Chaudhari"
        msg['To'] = email[0] + "," + email[1]

        with open("content.txt") as content:
            msg_content = content.read()
            msg.set_content(msg_content)

        with open("sample_output/marksheet/" + file, "rb") as f:
            file_data = f.read()
            msg.add_attachment(file_data, maintype="application", subtype="xlsx", filename=file)

        try:
            server.send_message(msg)
        except Exception as e:
            print("Error occurred: ", e)
        print("Email sent to ", email)

    def send_email(self):
        put_text("Sending Emails to students.......")
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login('danielisgamingtoday@gmail.com', "thisIsPassword@")
        for file in os.listdir("sample_output/marksheet"):
            if file == "concise_marksheet.csv":
                continue
            self.send_individual_mail(server, file, self.roll_to_email[file])
        put_success("Emails sent to concerned students")

    def marksheet_from_roll(self, roll, name, ans, stud_ans, total, neg_marks, pos, neg):
        wb = openpyxl.Workbook()
        total_qns = len(ans)
        img = Image("logo.png")
        sheet = wb.active
        sheet.add_image(img, "A1")
        sheet.title = "Quiz.marks"

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 25

        b_font = Font(bold=True,size=18)
        n_font = Font(size=14)
        green_font= Font(bold=True,size=14,color = '00008000')
        red_font= Font(bold=True,size=14,color = '00FF0000')
        blue_font= Font(bold=True,size=14,color = '000000FF')
        
        sheet.merge_cells('A6:E7')
        font = Font(size=17,bold=True,underline='single')
        cell = sheet.cell(row=6, column=1)  
        cell.value = "marksheet"
        cell.font = font 
        cell.alignment = Alignment(horizontal='center', vertical='center')  

        cell = sheet["A8"]
        cell.value = "Name"
        cell.font = n_font

        cell = sheet["B8"]
        cell.value = name
        cell.font = b_font

        cell = sheet["D8"]
        cell.value = "Exam"
        cell.font = n_font

        cell = sheet["E8"]
        cell.value = "Quiz"
        cell.font = n_font

        cell = sheet["A9"]
        cell.value = "Roll Number"
        cell.font = n_font

        cell = sheet["B9"]
        cell.value = roll
        cell.font = b_font

        double = Side(border_style='medium', color="00000000")
        border = Border(left=double,right=double,top=double,bottom=double)
        cell = sheet["A11"]
        cell.border = border

        cell = sheet["B11"]
        cell.value = "Right"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["C11"]
        cell.value = "Wrong"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["D11"]
        cell.value = "Not Attempt"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["E11"]
        cell.value = "max"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center') 

        cell = sheet["A14"]
        cell.border = border
        cell.value = "No."
        cell.font = b_font
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["B14"]
        cell.value = total/pos
        cell.font = green_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["C14"]
        cell.value = neg_marks/neg
        cell.font = red_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["D14"]
        cell.value = total_qns - (total/pos + neg_marks/neg)
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["E14"]
        cell.value = total_qns
        cell.font = blue_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center') 


        cell = sheet["A13"]
        cell.border = border
        cell.value = "marking"
        cell.font = b_font
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["B13"]
        cell.value = pos
        cell.font = green_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["C13"]
        cell.value = neg
        cell.font = red_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["D13"]
        cell.value = 0
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["E13"]
        cell.font = blue_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center') 


        cell = sheet["A14"]
        cell.border = border
        cell.value = "Total"
        cell.font = b_font
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["B14"]
        cell.value = total
        cell.font = green_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["C14"]
        cell.value = neg_marks
        cell.font = red_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["D14"]
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')  

        cell = sheet["E14"]
        cell.value = str(total+neg_marks) + "/" + str(pos*total_qns)
        cell.font = blue_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center') 

        cell = sheet["A17"]
        cell.value = "Student Ans"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center') 

        cell = sheet["B17"]
        cell.value = "Correct Ans"
        cell.font = b_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')


    def get_answer_map_from_roll(self, roll, df):
        ans = df.loc[df['Roll Number'] == roll]
        answers = ans.values.tolist()[0][7:]
        answer_map = {}
        c =1
        for ans in answers:
            answer_map[c] = ans
            c+=1 
        return answer_map


    def marksheets(self, csvreader, masreader, pos, neg):
        stdmpp = {}
        header = []
        for ind in masreader.index:
            stdmpp[masreader['roll'][ind]] = masreader['name'][ind]

        with open("sample_input/save_response.csv","w+") as csvfile:
            csvreader.to_csv("sample_input/save_response.csv")

        df = pd.read_csv("sample_input/save_response.csv")
        print(df)

        score_after_negative = []
        statusAns = []

        ans = self.get_answer_map_from_roll("ANSWER", df)
        print(ans)
        for ind in df.index:
            stud_ans = self.get_answer_map_from_roll(df['Roll Number'][ind], df) 
            self.roll_to_email[df['Roll Number'][ind] + ".xlsx"]=[df['Email address'][ind] , df["IITP webmail"][ind] ]
            total = 0
            actual_marks = 0
            neg_marks = 0
            for i in range(1,len(stud_ans)+1):
                if stud_ans[i] == ans[i]:
                    total += pos
                elif isinstance(stud_ans[i], float) and math.isnan(stud_ans[i]):
                    pass
                else:
                    neg_marks +=neg 
            actual_marks = total + neg_marks
            score_after_negative.append(str(actual_marks)+"/"+str(pos*len(stud_ans)))
            statusAns.append("["+str(int(total/pos))+","+str(int(neg_marks/neg))+","+str(int(len(stud_ans)-total/pos-neg_marks/neg))+"]")
            self.marksheet_from_roll(df['Roll Number'][ind], df['Name'][ind],ans,stud_ans,total,neg_marks,pos,neg)


        print(len(stud_ans))
        df.insert(loc=6, column='Score_After_Negative', value=score_after_negative)
        df.insert(loc=len(df.columns), column='statusAns', value=statusAns)
        df.to_csv('sample_output/marksheet/concise_marksheet.csv',index=False)


    def content_to_pandas(self, content):
        with open("tmp.csv", "w") as csv_file:
            writer = csv.writer(csv_file, delimiter=",")
            for line in content:
                writer.writerow(line.split(","))
        return pd.read_csv("tmp.csv")


    def main(self):
        master_roll = file_upload(label='Please select the master Roll file', accept=".csv", 
                                  name=None, placeholder='Choose file', multiple=False, max_size=0, max_total_size=0, required=True)
        master_roll_csv = self.content_to_pandas(master_roll['content'].decode('utf-8').splitlines())

        response_file = file_upload(label='Please select the Responses file', accept=".csv", name=None,
                                      placeholder='Choose file', multiple=False, max_size=0, max_total_size=0, required=True)
        response_file_csv = self.content_to_pandas(response_file['content'].decode('utf-8').splitlines())

        positive_marks = input("What is the positive marks per question", type=NUMBER)
        neg_marks = input("What is negative marks per question (write 0 if none)", type=NUMBER)

        self.marksheets(response_file_csv, master_roll_csv, positive_marks, neg_marks)

        put_success("Yayy!!! Your marksheets have been generated", closable=True)

        req_action = "generate pdf"
        while req_action != "none":
            req_action = actions(label="Please Select your action",
                                 buttons=[{'label': 'Download marksheets', 'value': "marksheets"},
                                          {'label': 'Send Email to students', 'value': "Email"},
                                          {'label': 'I wish to exit', 'value': "none"}])
            if req_action == "marksheets":
                put_html("<h2>Your marksheets have been downloaded, please check your sample_output folder</h2>")
            if req_action == "Email":
                self.send_email()
                put_html("<h2>The mail has been sent to the concerned students</h2>")

        put_success("This was my project, hope you liked it :)")



generator = QuizMarksheetGenerator()
start_server(generator.main, port=3001)
